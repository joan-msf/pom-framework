import os
from openpyxl import load_workbook
from src.utils.logger import get_logger


class ExcelReader:
    """
    ExcelReader class for reading spreadsheets containing test data.

    Example usage:
        reader = ExcelReader("path/to/test_data.xlsx")
        sheet_names = reader.get_sheet_names()
        data = reader.get_sheet_data("Sheet1")
        specific_value = reader.get_cell_value("Sheet1", row=2, column=3)
    """

    def __init__(self, file_path):
        """
        Initialize ExcelReader with the path to the Excel file.

        :param file_path: Path to the Excel (.xlsx) file.
        :raises FileNotFoundError: If the specified file does not exist.
        """
        self.logger = get_logger(self.__class__.__name__)
        self.file_path = file_path

        if not os.path.exists(self.file_path):
            self.logger.error(f"Excel file not found at {self.file_path}")
            raise FileNotFoundError(f"Excel file not found at {self.file_path}")

        try:
            self.workbook = load_workbook(filename=self.file_path, data_only=True)
            self.logger.info(f"Excel file loaded: {self.file_path}")
        except Exception as e:
            self.logger.error(f"Failed to load Excel file: {self.file_path}. Error: {str(e)}")
            raise e

    def get_sheet_names(self):
        """
        Retrieve all sheet names in the workbook.

        :return: List of sheet names.
        """
        return self.workbook.sheetnames

    def get_sheet_data(self, sheet_name):
        """
        Read data from a specified sheet and return it as a list of dictionaries.
        The first row is assumed to contain the header names.

        :param sheet_name: Name of the sheet to read.
        :return: List of dictionaries mapping header names to cell values.
        :raises ValueError: If the sheet name does not exist.
        """
        if sheet_name not in self.workbook.sheetnames:
            self.logger.error(f"Sheet '{sheet_name}' not found in {self.file_path}")
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook")

        sheet = self.workbook[sheet_name]
        data = []
        rows = list(sheet.rows)

        if not rows:
            self.logger.warning(f"Sheet '{sheet_name}' is empty")
            return data

        # Assume first row contains headers.
        headers = [cell.value for cell in rows[0]]

        for row in rows[1:]:
            row_data = {}
            for header, cell in zip(headers, row):
                row_data[header] = cell.value
            data.append(row_data)

        self.logger.info(f"Read {len(data)} rows from sheet '{sheet_name}'")
        return data

    def get_cell_value(self, sheet_name, row, column):
        """
        Retrieve the value of a specific cell in the specified sheet.

        :param sheet_name: Name of the sheet.
        :param row: Row number (1-indexed).
        :param column: Column number (1-indexed) or column letter (e.g., 'A').
        :return: The cell value.
        :raises ValueError: If the sheet name does not exist.
        """
        if sheet_name not in self.workbook.sheetnames:
            self.logger.error(f"Sheet '{sheet_name}' not found in {self.file_path}")
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook")

        sheet = self.workbook[sheet_name]
        # Handle column as either an integer index or a letter
        if isinstance(column, int):
            cell = sheet.cell(row=row, column=column)
        else:
            cell = sheet[f"{column}{row}"]

        cell_value = cell.value
        self.logger.info(f"Value at {sheet_name} (row {row}, column {column}): {cell_value}")
        return cell_value
